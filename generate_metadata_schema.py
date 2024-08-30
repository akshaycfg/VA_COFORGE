import os
from openpyxl import load_workbook
import json
import importlib
from config import cirium_otp




def get_metadata(project_name):
    try:
        with open(f'projects.json', 'r') as json_file:
            projects = json.load(json_file)
        list_of_configs=projects[project_name]['list_of_configs']

        for config_file in list_of_configs:
            module=importlib.import_module(f'config.{config_file[:-3]}')
            config=getattr(module, config_file[:-3])
            master_metadata = {
                "system": config["master_metadata_system"],
                "data_set_name": config["data_set_name"],
                "system_name": config["master_metadata_system_path"]

            }
            checklist_path=config['file_path']
            checklist = load_workbook(filename=checklist_path)
            tables = config['tables'].split(',')
            for sheet in checklist.sheetnames:
                if sheet in tables:
                    table_list = []
                    table = {
                        "table_name": sheet,
                        "model_type": config["model_type"],
                        "strategy": config["strategy"],
                        "name_unq_id": f"{sheet}_unq_id",
                        "name_grp_id": f"{sheet}_grp_id",
                        "name_child_grp_id": f"{sheet}_child_grp_id",
                        "name_updated_at": config["name_updated_at"],
                        "table_time_zone": config["table_time_zone"],
                        "timestamp_format": config["timestamp_format"],
                        "date_only_format": config["date_only_format"],
                        "invalidate_hard_deletes": config["invalidate_hard_deletes"],
                        "storageLevel": config["storageLevel"],
                        "silver_partition_start": config["silver_partition_start"],
                        "silver_partition_field": config["silver_partition_field"]
                    }
                    #print(table)
                    field_list = []
                    for r in checklist[sheet].iter_rows(values_only=True):
                        if r[config["bronze_to_silver_mapping_ix"]]=='One to One' and type(r[config["pos_ix"]]).__name__=="int":
                            #print(r[config["pos_ix"]])
                            field = {
                                "name": r[config["column_name_ix"]],
                                "source_data_type": "",
                                "parquet_data_type": "",
                                "bronze_data_type": r[config["bronze_data_type_ix"]].replace(',', ', ').lower(),
                                "silver_data_type": r[config["silver_data_type_ix"]].replace(',', ', '),
                                "ordinal_position": r[config["pos_ix"]],
                                "extract": "True",
                                "pii": "False",
                                "is_nullable": r[config["is_nullable_ix"]]
                            }
                            field_list.append(field)

                    primary_key_field_list = []
                    for r in checklist[sheet].iter_rows(values_only=True):
                        if r[config["bronze_to_silver_mapping_ix"]]=='One to One' and type(r[config["pos_ix"]]).__name__=="int" and (r[config["unique_key_ix"]] == "Y"):
                            primary_key_field = {
                                "name": r[config["column_name_ix"]]
                            }
                            primary_key_field_list.append(primary_key_field)

                    group_key_field_list = []
                    for r in checklist[sheet].iter_rows(values_only=True):
                        if r[config["bronze_to_silver_mapping_ix"]] == 'One to One' and type(r[config["pos_ix"]]).__name__ == "int" and (r[config["soft_delete_key_ix"]] == "Y-1"):
                            group_key_field = {
                                "name": r[config["column_name_ix"]]
                            }
                            group_key_field_list.append(group_key_field)

                    child_group_key_field_list = []
                    for r in checklist[sheet].iter_rows(values_only=True):
                        if r[config["bronze_to_silver_mapping_ix"]] == 'One to One' and type(r[config["pos_ix"]]).__name__ == "int" and (r[config["soft_delete_key_ix"]] == "Y-2"):
                            child_group_key_field = {
                                "name": r[config["column_name_ix"]]
                            }
                            child_group_key_field_list.append(child_group_key_field)
                    duplicate_ranking_field_list = []
                    for r in checklist[sheet].iter_rows(values_only=True):
                        if r[config["bronze_to_silver_mapping_ix"]] == 'One to One' and type(r[config["pos_ix"]]).__name__ == "int" and (r[config["duplicate_ranking_ix"]] == "Y"):
                            duplicate_ranking_field = {
                                "name": r[config["column_name_ix"]]
                            }
                            duplicate_ranking_field_list.append(duplicate_ranking_field)

                    schema_value = {
                        "fields": field_list,
                        "primary_key_fields": primary_key_field_list,
                        "group_key_fields": group_key_field_list,
                        "child_group_key_fields": child_group_key_field_list,
                        "duplicate_ranking_fields": duplicate_ranking_field_list
                    }

                    schema = {
                        "schema": schema_value
                    }
                    table.update(schema)
                    table_list.append(table)
                    tables = {"tables": table_list}
                    master_metadata.update(tables)
                    if not os.path.exists(f"{project_name}/{config_file[:-3]}"):
                        os.mkdir(f"{project_name}/{config_file[:-3]}")
                    with open(f"{project_name}/{config_file[:-3]}/metadata.json", "w") as outfile:
                        json.dump(master_metadata, outfile, indent=4)
        return {"status": "Metadata generated Successfully"}
    except Exception as e:
        er_msg=f"Error in execution: {e}"
        print(er_msg)
        return {"status": er_msg}



def get_inventory(project_name):
    try:
        with open(f'projects.json', 'r') as json_file:
            projects = json.load(json_file)
        list_of_configs=projects[project_name]['list_of_configs']

        inventory={}

        for config_file in list_of_configs:
            module=importlib.import_module(f'config.{config_file[:-3]}')
            config=getattr(module, config_file[:-3])
            checklist_path=config['file_path']
            checklist = load_workbook(filename=checklist_path)
            tables = config['tables'].split(',')
            inventory[config_file] = tables
        with open(f"{project_name}/inventory.json", "w") as outfile:
            json.dump(inventory, outfile, indent=4)


        return inventory
    except Exception as e:
        er_msg=f"Error in execution: {e}"
        print(er_msg)
        return {"status": er_msg}

